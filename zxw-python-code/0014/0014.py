#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''第 0014 题： 纯文本文件 student.txt为学生信息, 里面的内容（包括花括号）如下所示：
{
    "1":["张三",150,120,100],
    "2":["李四",90,99,95],
    "3":["王五",60,66,68]
}
请将上述内容写到 student.xls 文件中。'''





from openpyxl import Workbook
import json

def txt_to_xls(filename):
    file = open(filename,'r', encoding='utf-8')
    file_cintent = json.load(file, encoding='utf-8')
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Students'
    for i in range(1,len(file_cintent)+1):
        sheet.cell(row=i, column = 1).value = i
        print(file_cintent)
        list = file_cintent.get(str(i))
        j = 2
        for cell in list:
            sheet.cell(row=i, column=j).value = cell
            j += 1


    wb.save('student.xlsx')


txt_to_xls('student.txt')
