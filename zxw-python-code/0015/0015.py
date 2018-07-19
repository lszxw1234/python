#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''第 0015 题： 纯文本文件 city.txt为城市信息, 里面的内容（包括花括号）如下所示：
{
    "1" : "上海",
    "2" : "北京",
    "3" : "成都"
}
请将上述内容写到 city.xls 文件中。'''

from openpyxl import Workbook
import json

def txt_to_xls(filename):
    file = open(filename,'r', encoding='utf-8')
    file_cintent = json.load(file, encoding='utf-8')
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Citys'
    for i in range(1,len(file_cintent)+1):
        sheet.cell(row=i, column = 1).value = i
        print(file_cintent)
        list = file_cintent.get(str(i))
        j = 2
        if callable(list):
            for cell in list:
                sheet.cell(row=i, column=j).value = cell
                j += 1
        else:
            sheet.cell(row=i, column=2).value = list


    wb.save('city.xlsx')


txt_to_xls('city.txt')
